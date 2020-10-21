#coding=utf-8
import time
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import os
class ExcelUtil:
    def __init__(self,excel_path=None):
        path=os.path.abspath(os.path.join(os.getcwd(),".."))+"\config\keyword.xlsx"
        if excel_path == None:
            #self.excel_path = r"G:\pythonselenium3lianxi\shuju2\config\keyword.xls"
            self.excel_path = path
        else:
            self.excel_path = excel_path




    #传入参数，打开表格
    def get_data(self):
        data = load_workbook(self.excel_path)
        tables = data['Sheet1']
        return tables
    #获取行数
    def get_lines(self):
        rows = self.get_data().max_row
        if rows >=1:
            return rows
        return None


    #获取单元格的数据
    def get_col_value(self,row,col):
        if self.get_lines() > row:
            data = self.get_data().cell(row,col).value
            return data
        return None
    #写入数据
    def write_value(self,row,col,value):
        read_value = load_workbook(self.excel_path)
        sheet_data = read_value['Sheet1']
        sheet_data.cell(row,col,value)
        read_value.save(self.excel_path)
        time.sleep(1)
if __name__ == '__main__':
    a=ExcelUtil().get_col_value(7,7)
    print(a)
